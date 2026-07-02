"""
Excel Export Utilities for Employee Data
Handles bulk export with customizable columns and filters
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
from io import BytesIO

# Define all available export columns
EXPORT_COLUMNS = {
    'basic': {
        'label': 'Basic Information',
        'fields': [
            ('employee_code', 'Employee Code'),
            ('first_name', 'First Name'),
            ('middle_name', 'Middle Name'),
            ('last_name', 'Last Name'),
            ('full_name', 'Full Name'),
            ('date_of_birth', 'Date of Birth'),
            ('gender', 'Gender'),
            ('blood_group', 'Blood Group'),
            ('marital_status', 'Marital Status'),
        ]
    },
    'contact': {
        'label': 'Contact Information',
        'fields': [
            ('mobile_number', 'Mobile Number'),
            ('whatsapp_number', 'WhatsApp Number'),
            ('personal_email', 'Personal Email'),
        ]
    },
    'employment': {
        'label': 'Employment Details',
        'fields': [
            ('designation', 'Designation'),
            ('department', 'Department'),
            ('employment_type', 'Employment Type'),
            ('joining_date', 'Date of Joining'),
            ('relieving_date', 'Relieving Date'),
            ('employee_status', 'Status'),
            ('manager_name', 'Reporting Manager'),
        ]
    },
    'address': {
        'label': 'Address',
        'fields': [
            ('perm_address', 'Permanent Address'),
            ('perm_city', 'Permanent City'),
            ('perm_state', 'Permanent State'),
            ('perm_postal_code', 'Permanent Postal Code'),
            ('curr_address', 'Current Address'),
            ('curr_city', 'Current City'),
            ('curr_state', 'Current State'),
            ('curr_postal_code', 'Current Postal Code'),
        ]
    },
    'bank': {
        'label': 'Bank & PF Details',
        'fields': [
            ('bank_name', 'Bank Name'),
            ('account_number', 'Account Number'),
            ('branch_name', 'Branch Name'),
            ('ifsc_code', 'IFSC Code'),
            ('pf_number', 'PF Number'),
            ('uan_number', 'UAN Number'),
            ('esi_number', 'ESI Number'),
        ]
    },
    'family': {
        'label': 'Family Information',
        'fields': [
            ('mother_name', "Mother's Name"),
            ('father_name', "Father's Name"),
            ('spouse_name', 'Spouse Name'),
            ('epf_member', 'Member of Provident Fund (EPF Act)'),
            ('family_pension_scheme', 'Covered by Family Pension Scheme (EPF Act)'),
        ]
    },
    'asset': {
        'label': 'Asset Details',
        'fields': [
            ('access_office_1', 'Access Office Reference 1'),
            ('access_office_2', 'Access Office Reference 2'),
            ('access_office_3', 'Access Office Reference 3'),
            ('asset_phone', 'Asset Phone Number'),
            ('bike_card', 'Bike Access Card Number'),
        ]
    },
    'other': {
        'label': 'Other Details',
        'fields': [
            ('nationality', 'Nationality'),
            ('religion', 'Religion'),
            ('place_of_birth', 'Place of Birth'),
            ('mother_tongue', 'Mother Tongue'),
        ]
    },
}


def get_employee_data_row(employee, selected_fields):
    """Extract data for selected fields from employee object"""
    row = []

    for field_key in selected_fields:
        value = None

        # Basic fields
        if field_key == 'full_name':
            value = f"{employee.first_name or ''} {employee.middle_name or ''} {employee.last_name or ''}".strip()
        elif field_key == 'manager_name':
            value = f"{employee.manager.first_name} {employee.manager.last_name}" if employee.manager else ''

        # Address fields
        elif field_key == 'perm_address':
            parts = [employee.perm_address_line1, employee.perm_address_line2]
            value = ', '.join([p for p in parts if p])
        elif field_key == 'curr_address':
            parts = [employee.curr_address_line1, employee.curr_address_line2]
            value = ', '.join([p for p in parts if p])

        # Family fields
        elif field_key == 'mother_name':
            value = employee.family_info.mother_name if employee.family_info else ''
        elif field_key == 'father_name':
            value = employee.family_info.father_name if employee.family_info else ''
        elif field_key == 'spouse_name':
            value = employee.family_info.spouse_name if employee.family_info else ''
        elif field_key == 'epf_member':
            value = 'Yes' if (employee.family_info and employee.family_info.epf_member) else 'No'
        elif field_key == 'family_pension_scheme':
            if employee.family_info and employee.family_info.family_pension_scheme:
                value = 'Yes' if employee.family_info.family_pension_scheme == 'yes' else 'No'
            else:
                value = 'No'

        # Asset fields
        elif field_key == 'access_office_1':
            value = employee.asset_info.access_office_field_1 if employee.asset_info else ''
        elif field_key == 'access_office_2':
            value = employee.asset_info.access_office_field_2 if employee.asset_info else ''
        elif field_key == 'access_office_3':
            value = employee.asset_info.access_office_field_3 if employee.asset_info else ''
        elif field_key == 'asset_phone':
            value = employee.asset_info.phone_number if employee.asset_info else ''
        elif field_key == 'bike_card':
            value = employee.asset_info.bike_access_card_number if employee.asset_info else ''

        # Direct attribute access for standard fields
        elif hasattr(employee, field_key):
            value = getattr(employee, field_key)

            # Format dates
            if value and hasattr(value, 'strftime'):
                value = value.strftime('%d-%m-%Y')

        row.append(value or '')

    return row


def generate_employee_excel(employees, selected_fields, filters_applied=None):
    """
    Generate Excel file with employee data

    Args:
        employees: List of Employee objects
        selected_fields: List of field keys to include
        filters_applied: Dict describing applied filters (for metadata)

    Returns:
        BytesIO object containing Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Data"

    # Get field labels for selected fields
    field_labels = []
    field_keys = []
    for category, data in EXPORT_COLUMNS.items():
        for field_key, field_label in data['fields']:
            if field_key in selected_fields:
                field_labels.append(field_label)
                field_keys.append(field_key)

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # Write metadata
    ws['A1'] = 'Employee Data Export'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f'Generated on: {datetime.now().strftime("%d-%m-%Y %H:%M:%S")}'
    ws['A3'] = f'Total Records: {len(employees)}'

    if filters_applied:
        row_num = 4
        for filter_name, filter_value in filters_applied.items():
            ws[f'A{row_num}'] = f'{filter_name}: {filter_value}'
            row_num += 1

    # Start data table at row 6
    header_row = 6

    # Write headers
    for col_idx, label in enumerate(field_labels, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border

    # Write employee data
    for row_idx, employee in enumerate(employees, start=header_row + 1):
        row_data = get_employee_data_row(employee, field_keys)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center")

    # Auto-adjust column widths
    for col_idx in range(1, len(field_labels) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0

        for row in ws[column_letter]:
            if row.value:
                max_length = max(max_length, len(str(row.value)))

        adjusted_width = min(max_length + 2, 50)  # Max 50 chars width
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = f'A{header_row + 1}'

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def get_all_field_keys():
    """Get list of all available field keys"""
    all_fields = []
    for category, data in EXPORT_COLUMNS.items():
        all_fields.extend([field[0] for field in data['fields']])
    return all_fields
